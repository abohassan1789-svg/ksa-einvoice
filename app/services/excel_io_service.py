"""Reusable Excel template export / import engine for the CRM review screens.

One engine, driven entirely by per-screen metadata, so no screen duplicates any
import/export code. A screen only declares (via :data:`IMPORT_RULES` and its
existing :class:`~app.services.review_data_service.TableSpec`):

* which grid columns appear in the template (taken from ``TableSpec.list_columns``),
* which of them are required,
* each column's data type,
* the unique/duplicate signature used to reject duplicate rows.

The public surface is :class:`ExcelTemplateService`:

* :meth:`~ExcelTemplateService.export_template` writes an empty, professionally
  formatted ``.xlsx`` with exactly the screen's visible columns.
* :meth:`~ExcelTemplateService.build_import_rows` reads a filled workbook,
  validates every row (required fields, data types, foreign-key resolution and
  duplicates) and returns per-row results for a review window.
* :meth:`~ExcelTemplateService.save_valid_rows` persists only the valid rows,
  reusing ``ReviewDataService.save_record`` so all existing coercion / id
  allocation rules apply unchanged.

The row-level validation (:func:`build_import_rows`) is deliberately split from
workbook I/O so it can be unit-tested with plain Python lists and a fake data
service, without touching a database or the filesystem.
"""

from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.review_data_service import TABLE_SPECS, TableSpec
from app.ui.common.theme import EXTRA_COLUMN_LABELS

# Screen keys that support Excel template export/import. Kept explicit so the
# feature stays scoped to exactly the five requested CRM screens.
SUPPORTED_KEYS: tuple[str, ...] = (
    "customers",
    "products",
    "companies",
    "places",
    "case_statuses",
    "employees",
    "daily_followups",
)


@dataclass(frozen=True)
class ExcelColumn:
    """One column of a screen's Excel template."""

    key: str
    header: str
    data_type: str = "text"  # text | int | float | date
    required: bool = False
    # Primary key: auto-generated on insert, so it is exported (it is visible in
    # the grid) but ignored when importing.
    id_field: bool = False
    # Shown in the template for context but never written back (e.g. the
    # customer name column on the daily follow-up grid, resolved via phone).
    display_only: bool = False
    # Database field this column populates. Defaults to ``key``.
    target_field: str = ""
    # Foreign-key resolver name (see ExcelTemplateService._RESOLVERS), e.g.
    # turning a phone number into a customer_id.
    resolver: str = ""

    @property
    def target(self) -> str:
        return self.target_field or self.key


@dataclass(frozen=True)
class ExcelIoSpec:
    """Everything the engine needs to export/import one screen."""

    key: str
    table_spec: TableSpec
    sheet_title: str
    columns: tuple[ExcelColumn, ...]
    unique_key: tuple[str, ...]  # target-field names forming the duplicate signature
    unique_label: str

    @property
    def headers(self) -> list[str]:
        return [column.header for column in self.columns]


@dataclass(frozen=True)
class ImportRules:
    """Per-screen import configuration (the only thing a screen must declare)."""

    required: frozenset[str] = frozenset()
    unique_key: tuple[str, ...] = ()
    unique_label: str = ""
    resolvers: Mapping[str, str] = field(default_factory=dict)  # column key -> resolver
    targets: Mapping[str, str] = field(default_factory=dict)  # column key -> db field
    data_types: Mapping[str, str] = field(default_factory=dict)  # column key -> dtype
    display_only: frozenset[str] = frozenset()


# --- Per-screen rules -------------------------------------------------------
# Columns themselves come from each TableSpec.list_columns (the exact visible
# grid columns, same names and order). These rules add only what the grid
# metadata cannot express: which columns are required for import, the duplicate
# signature, and (for daily follow-ups) foreign-key resolvers.
IMPORT_RULES: dict[str, ImportRules] = {
    "customers": ImportRules(
        required=frozenset({"customer_name", "phone_number"}),
        unique_key=("phone_number",),
        unique_label="رقم الهاتف",
    ),
    "places": ImportRules(
        required=frozenset({"place_number"}),
        unique_key=("place_number",),
        unique_label="رقم / اسم المنطقة",
    ),
    "case_statuses": ImportRules(
        required=frozenset({"case_status_name"}),
        unique_key=("case_status_name",),
        unique_label="اسم الحالة",
    ),
    "products": ImportRules(
        # item_code is auto-assigned (id_field) and total is a generated column
        # (display_only) — both appear in the template but are never imported.
        # Duplicate items are detected by name (اسم الصنف), per the chosen rule.
        required=frozenset({"item_name"}),
        unique_key=("item_name",),
        unique_label="اسم الصنف",
        display_only=frozenset({"total"}),
    ),
    "companies": ImportRules(
        # id is auto-assigned (id_field) — it appears in the template but is never
        # imported. name_ar / commercial_registration / vat_number are required.
        # Duplicates are previewed by commercial registration (السجل التجاري); the
        # database UNIQUE indexes on registration AND VAT are the final guard on
        # save (a duplicate VAT slips past the preview but is rejected at save and
        # reported per-row).
        required=frozenset({"name_ar", "commercial_registration", "vat_number"}),
        unique_key=("commercial_registration",),
        unique_label="السجل التجاري",
    ),
    "employees": ImportRules(
        required=frozenset({"employee_name", "phone_number"}),
        unique_key=("phone_number",),
        unique_label="رقم الهاتف",
    ),
    "daily_followups": ImportRules(
        required=frozenset({"customer_phone", "follow_up_date"}),
        unique_key=("customer_id", "follow_up_date", "case_status_id"),
        unique_label="العميل + تاريخ المتابعة + الحالة",
        resolvers={
            "customer_phone": "customer_by_phone",
            "case_status_name": "case_status_by_name",
        },
        targets={
            "customer_phone": "customer_id",
            "case_status_name": "case_status_id",
        },
        data_types={"follow_up_date": "date"},
        display_only=frozenset({"customer_name"}),
    ),
}


def _header_for(spec: TableSpec, column_key: str) -> str:
    for field_spec in spec.fields:
        if field_spec.name == column_key:
            return field_spec.label
    return EXTRA_COLUMN_LABELS.get(column_key, column_key)


def _data_type_for(spec: TableSpec, column_key: str, rules: ImportRules) -> str:
    if column_key in rules.data_types:
        return rules.data_types[column_key]
    for field_spec in spec.fields:
        if field_spec.name == column_key:
            return field_spec.data_type
    return "text"


def build_io_spec(spec: TableSpec) -> ExcelIoSpec:
    """Derive the Excel import/export spec from a screen's ``TableSpec``.

    Columns mirror the visible grid (``list_columns``) exactly — same names and
    order — with headers, data types and required flags pulled from the field
    metadata plus the per-screen :data:`IMPORT_RULES`.
    """
    rules = IMPORT_RULES.get(spec.key, ImportRules())
    columns: list[ExcelColumn] = []
    for column_key in spec.list_columns:
        header = _header_for(spec, column_key)
        is_id = column_key == spec.primary_key
        columns.append(
            ExcelColumn(
                key=column_key,
                header=header,
                data_type="int" if is_id else _data_type_for(spec, column_key, rules),
                required=column_key in rules.required,
                id_field=is_id,
                display_only=column_key in rules.display_only,
                target_field=rules.targets.get(column_key, ""),
                resolver=rules.resolvers.get(column_key, ""),
            )
        )
    return ExcelIoSpec(
        key=spec.key,
        table_spec=spec,
        sheet_title=spec.title,
        columns=tuple(columns),
        unique_key=rules.unique_key,
        unique_label=rules.unique_label,
    )


def io_spec_for(key: str) -> ExcelIoSpec:
    if key not in SUPPORTED_KEYS:
        raise KeyError(f"Excel import/export is not supported for screen: {key}")
    return build_io_spec(TABLE_SPECS[key])


@dataclass
class ImportRow:
    """Result of validating one imported data row."""

    row_number: int  # 1-based data row number (Excel row = row_number + 1)
    values: dict[str, str]  # header -> raw display text
    payload: dict[str, Any] = field(default_factory=dict)  # db field -> coerced value
    errors: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors

    @property
    def status(self) -> str:
        return "valid" if self.is_valid else "error"


class ExcelWorkbookError(Exception):
    """Raised when a workbook cannot be read or its columns do not match."""


class ExcelTemplateService:
    """Reusable export/import engine shared by every supported screen."""

    def __init__(self, data_service: Any) -> None:
        # ``data_service`` is a ReviewDataService (or any object exposing the few
        # methods used below), injected so the engine stays DB-agnostic and
        # unit-testable with a fake.
        self.data_service = data_service
        self._RESOLVERS: dict[str, Callable[[str], Any]] = {
            "customer_by_phone": self._resolve_customer_by_phone,
            "case_status_by_name": self._resolve_case_status_by_name,
        }

    # -- Export --------------------------------------------------------------

    def export_template(self, io_spec: ExcelIoSpec, path: str | Path) -> Path:
        """Write an empty, formatted template with the screen's visible columns.

        Formatting: bold header row, frozen top row, right-to-left sheet (Arabic)
        and auto-fit column widths.
        """
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = _safe_sheet_title(io_spec.sheet_title)
        worksheet.sheet_view.rightToLeft = True

        headers = io_spec.headers
        worksheet.append(headers)

        header_fill = PatternFill("solid", fgColor="137A38")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Freeze the header row so it stays visible while filling many rows.
        worksheet.freeze_panes = "A2"

        for index, header in enumerate(headers, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(str(header)) + 6, 14), 45
            )

        workbook.save(output_path)
        return output_path

    # -- Import: read + validate --------------------------------------------

    def read_workbook(self, io_spec: ExcelIoSpec, path: str | Path) -> list[ImportRow]:
        """Read a filled template, validate its columns, and validate every row.

        Raises :class:`ExcelWorkbookError` if the file cannot be opened or its
        header row does not match the expected template columns.
        """
        header, data_rows = self._read_sheet(path)
        self.check_headers(io_spec, header)
        return self.build_import_rows(io_spec, data_rows)

    def _read_sheet(self, path: str | Path) -> tuple[list[str], list[list[Any]]]:
        try:
            workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
        except Exception as exc:  # openpyxl raises a variety of errors
            raise ExcelWorkbookError(f"تعذر فتح ملف الإكسل: {exc}") from exc
        try:
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            raise ExcelWorkbookError("ملف الإكسل فارغ.")
        header = ["" if cell is None else str(cell).strip() for cell in rows[0]]
        data_rows = [list(row) for row in rows[1:]]
        return header, data_rows

    def check_headers(self, io_spec: ExcelIoSpec, header: list[str]) -> None:
        """Ensure the workbook's columns match the template exactly (name + order)."""
        expected = io_spec.headers
        # Trim trailing empty header cells that spreadsheet apps sometimes add.
        trimmed = list(header)
        while trimmed and trimmed[-1] in ("", None):
            trimmed.pop()
        if trimmed == expected:
            return
        expected_set = set(expected)
        actual_set = set(trimmed)
        missing = [h for h in expected if h not in actual_set]
        extra = [h for h in trimmed if h not in expected_set]
        parts = ["أعمدة الملف لا تطابق قالب هذه الشاشة."]
        if missing:
            parts.append("أعمدة ناقصة: " + "، ".join(missing))
        if extra:
            parts.append("أعمدة غير متوقعة: " + "، ".join(extra))
        if not missing and not extra:
            parts.append("ترتيب الأعمدة غير صحيح.")
        parts.append("المتوقع: " + "، ".join(expected))
        raise ExcelWorkbookError("\n".join(parts))

    def build_import_rows(
        self, io_spec: ExcelIoSpec, data_rows: list[list[Any]]
    ) -> list[ImportRow]:
        """Validate raw data rows (columns already confirmed to match).

        Each row is checked for required fields, data-type correctness, foreign
        keys that resolve, and duplicates — both against existing database rows
        and against earlier rows in the same file.
        """
        existing = self._existing_signatures(io_spec)
        seen: dict[tuple, int] = {}
        results: list[ImportRow] = []
        for offset, raw in enumerate(data_rows, start=1):
            if _is_blank_row(raw):
                continue
            row = self._validate_row(io_spec, offset, raw)
            if row.is_valid and io_spec.unique_key:
                signature = self._signature(io_spec, row.payload)
                if signature in existing:
                    row.errors.append(f"قيمة مكررة موجودة مسبقًا ({io_spec.unique_label}).")
                elif signature in seen:
                    row.errors.append(
                        f"قيمة مكررة داخل الملف مع الصف رقم {seen[signature]} ({io_spec.unique_label})."
                    )
                else:
                    seen[signature] = offset
            results.append(row)
        return results

    def _validate_row(
        self, io_spec: ExcelIoSpec, row_number: int, raw: list[Any]
    ) -> ImportRow:
        values: dict[str, str] = {}
        payload: dict[str, Any] = {}
        errors: list[str] = []
        for index, column in enumerate(io_spec.columns):
            raw_cell = raw[index] if index < len(raw) else None
            text = _cell_text(raw_cell)
            values[column.header] = text
            if column.id_field or column.display_only:
                continue
            if text == "":
                if column.required:
                    errors.append(f"حقل مطلوب فارغ: {column.header}.")
                continue
            if column.resolver:
                resolved, error = self._resolve(column, text)
                if error:
                    errors.append(error)
                else:
                    payload[column.target] = resolved
                continue
            coerced, error = _coerce(column, text)
            if error:
                errors.append(error)
            else:
                payload[column.target] = coerced
        return ImportRow(row_number=row_number, values=values, payload=payload, errors=errors)

    # -- Import: save --------------------------------------------------------

    def save_valid_rows(
        self, io_spec: ExcelIoSpec, rows: list[ImportRow]
    ) -> tuple[int, list[tuple[int, str]]]:
        """Persist only the valid rows. Returns (saved_count, failures).

        Reuses ``ReviewDataService.save_record`` so id allocation, coercion and
        required-field checks stay identical to manual entry. Invalid rows are
        skipped entirely.
        """
        saved = 0
        failures: list[tuple[int, str]] = []
        for row in rows:
            if not row.is_valid:
                continue
            try:
                self.data_service.save_record(io_spec.table_spec, dict(row.payload), None)
                saved += 1
            except Exception as exc:  # keep going; report the row that failed
                message = str(exc)
                row.errors.append(message)
                failures.append((row.row_number, message))
        return saved, failures

    # -- Resolvers / duplicate helpers --------------------------------------

    def _resolve(self, column: ExcelColumn, text: str) -> tuple[Any, str | None]:
        resolver = self._RESOLVERS.get(column.resolver)
        if resolver is None:
            return None, f"لا يوجد محول للعمود: {column.header}."
        value = resolver(text)
        if value is None:
            return None, f"قيمة غير معروفة في {column.header}: {text}."
        return value, None

    def _resolve_customer_by_phone(self, text: str) -> Any:
        record = self.data_service.find_customer_by_phone(text)
        return record.get("customer_id") if record else None

    def _resolve_case_status_by_name(self, text: str) -> Any:
        record = self.data_service.find_case_status_by_name(text)
        return record.get("case_status_id") if record else None

    def _existing_signatures(self, io_spec: ExcelIoSpec) -> set[tuple]:
        if not io_spec.unique_key:
            return set()
        dtypes = self._unique_dtypes(io_spec)
        rows = self.data_service.fetch_column_values(
            io_spec.table_spec.table_name, list(io_spec.unique_key)
        )
        signatures: set[tuple] = set()
        for record in rows:
            signatures.add(
                tuple(
                    _normalize_signature(record.get(target), dtypes[target])
                    for target in io_spec.unique_key
                )
            )
        return signatures

    def _signature(self, io_spec: ExcelIoSpec, payload: dict[str, Any]) -> tuple:
        dtypes = self._unique_dtypes(io_spec)
        return tuple(
            _normalize_signature(payload.get(target), dtypes[target])
            for target in io_spec.unique_key
        )

    def _unique_dtypes(self, io_spec: ExcelIoSpec) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for target in io_spec.unique_key:
            dtype = "text"
            for column in io_spec.columns:
                if column.target == target:
                    dtype = "int" if column.resolver else column.data_type
                    break
            mapping[target] = dtype
        return mapping


# --- Module-level helpers ---------------------------------------------------


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        # Cells typed as dates come back as datetime; show only the date part.
        if value.time() == datetime.time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _coerce(column: ExcelColumn, text: str) -> tuple[Any, str | None]:
    """Coerce a non-empty cell to the column's data type."""
    if column.data_type == "int":
        try:
            return int(float(text)) if _looks_numeric(text) else int(text), None
        except (TypeError, ValueError):
            return None, f"قيمة رقمية غير صحيحة في {column.header}: {text}."
    if column.data_type == "float":
        try:
            return float(text), None
        except (TypeError, ValueError):
            return None, f"قيمة رقمية غير صحيحة في {column.header}: {text}."
    if column.data_type == "date":
        parsed = _parse_date(text)
        if parsed is None:
            return None, f"تاريخ غير صحيح في {column.header} (الصيغة YYYY-MM-DD): {text}."
        return parsed, None
    return text, None


def _looks_numeric(text: str) -> bool:
    try:
        float(text)
        return True
    except (TypeError, ValueError):
        return False


def _parse_date(text: str) -> datetime.date | None:
    candidate = text.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(candidate[:10], fmt).date()
        except ValueError:
            continue
    return None


def _normalize_signature(value: Any, dtype: str) -> Any:
    if value is None or value == "":
        return None
    if dtype == "int":
        try:
            return int(float(str(value).strip()))
        except (TypeError, ValueError):
            return str(value).strip()
    if dtype == "float":
        try:
            return float(str(value).strip())
        except (TypeError, ValueError):
            return str(value).strip()
    if dtype == "date":
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.isoformat()[:10] if not isinstance(value, datetime.datetime) else value.date().isoformat()
        return str(value).strip()[:10]
    return str(value).strip().casefold()


def _is_blank_row(raw: list[Any]) -> bool:
    return all(_cell_text(cell) == "" for cell in raw)


def _safe_sheet_title(title: str) -> str:
    invalid = set(r'[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in title).strip()
    return (cleaned or "Sheet")[:31]
