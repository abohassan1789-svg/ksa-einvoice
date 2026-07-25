import json

from openpyxl import load_workbook
import pytest

from app.reports import pdf_exporter as pdf_exporter_module
from app.reports.excel_exporter import ExcelExporter
from app.reports.pdf_exporter import PdfExporter
from app.reports.print_manager import PrintManager
from app.services.daily_followup_report_service import (
    DailyFollowupReportRequest,
    DailyFollowupReportService,
    DailyFollowupReportSettingsService,
    SmartFilterCondition,
)


class _Cursor:
    def __init__(self, rows):
        self._rows = rows

    def __iter__(self):
        return iter(self._rows)


class _Connection:
    def __init__(self, rows):
        self.rows = rows
        self.calls = []

    def execute(self, query, params=None):
        self.calls.append((query, params))
        return _Cursor(self.rows)


class _ConnectContext:
    def __init__(self, connection):
        self.connection = connection

    def __enter__(self):
        return self.connection

    def __exit__(self, *_args):
        return False


class _ReviewService:
    def __init__(self, rows=None):
        self.connection = _Connection(rows or [])
        self.customer_search_calls = []

    def connect(self):
        return _ConnectContext(self.connection)

    def list_case_statuses_for_selection(self):
        return [{"case_status_id": 1, "case_status_name": "Interested"}]

    def list_employees_for_selection(self):
        return [{"employee_id": 7, "employee_name": "Ahmed"}]

    def list_places_for_selection(self):
        return [{"place_id": 3, "place_number": "Cairo"}]

    def search_customers(self, keyword, limit=100):
        self.customer_search_calls.append((keyword, limit))
        return [{"customer_id": 1001, "customer_name": "Ali", "phone_number": "010"}]


def test_report_columns_are_discovered_from_daily_followup_screen_metadata():
    service = DailyFollowupReportService()

    keys = [column.key for column in service.columns]

    assert keys[:4] == [
        "daily_followup_id",
        "follow_up_date",
        "customer_id",
        "customer_name",
    ]
    assert "customer_phone" in keys
    assert "employee_name" in keys
    assert "case_status_name" in keys
    assert "notes" in keys
    assert "contact_count" not in keys


def test_build_query_rejects_unapproved_selected_columns():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(selected_columns=["notes; DROP TABLE customers"])

    with pytest.raises(ValueError, match="Unsupported report column"):
        service.build_query(request)


def test_build_query_uses_date_range_and_parameterized_and_filters():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(
        selected_columns=["daily_followup_id", "follow_up_date", "customer_name"],
        date_from="2026-06-01",
        date_to="2026-06-30",
        join_mode="AND",
        filters=[
            SmartFilterCondition("case_status_name", "equals", "Interested"),
            SmartFilterCondition("customer_name", "contains", "Mohamed"),
        ],
    )

    query, params = service.build_query(request)
    sql = str(query)

    assert "d.follow_up_date >= %s" in sql
    assert "d.follow_up_date < (%s::date + INTERVAL '1 day')" in sql
    assert "CAST(cs.case_status_name AS text) = %s" in sql
    assert "c.customer_name ILIKE %s" in sql
    assert " OR " not in sql.split(" WHERE ", 1)[1].split(" ORDER BY ", 1)[0]
    assert params == ["2026-06-01", "2026-06-30", "Interested", "%Mohamed%"]


def test_build_query_without_explicit_limit_returns_all_daily_followup_rows():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(selected_columns=["daily_followup_id", "customer_name"])

    query, params = service.build_query(request)

    assert "LIMIT" not in str(query)
    assert params == []


def test_build_query_keeps_explicit_limit_when_requested():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(
        selected_columns=["daily_followup_id"],
        limit=100,
    )

    query, params = service.build_query(request)

    assert "LIMIT %s" in str(query)
    assert params == [100]


def test_build_query_uses_or_join_mode_for_smart_filters_only():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(
        selected_columns=["daily_followup_id"],
        date_from="2026-06-01",
        join_mode="OR",
        filters=[
            SmartFilterCondition("employee_name", "starts_with", "Ahmed"),
            SmartFilterCondition("area_number", "ends_with", "12"),
        ],
    )

    query, params = service.build_query(request)
    sql = str(query)

    assert "d.follow_up_date >= %s" in sql
    assert "(e.employee_name ILIKE %s OR c.area_number ILIKE %s)" in sql
    assert params == ["2026-06-01", "Ahmed%", "%12"]


def test_build_query_casts_non_text_columns_for_text_match_operators():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(
        selected_columns=["customer_id"],
        filters=[SmartFilterCondition("customer_id", "contains", "1001")],
    )

    query, params = service.build_query(request)

    assert "CAST(d.customer_id AS text) ILIKE %s" in str(query)
    assert params == ["%1001%"]


def test_build_query_casts_text_columns_for_equals_to_avoid_type_mismatch():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(
        selected_columns=["employee_name"],
        filters=[SmartFilterCondition("employee_name", "equals", 7)],
    )

    query, params = service.build_query(request)

    assert "CAST(e.employee_name AS text) = %s" in str(query)
    assert params == ["7"]


def test_settings_service_persists_selected_columns_and_filters(tmp_path):
    path = tmp_path / "daily_followup_report_settings.json"
    service = DailyFollowupReportSettingsService(path)
    settings = {
        "selected_columns": ["follow_up_date", "customer_name"],
        "join_mode": "OR",
        "filters": [{"column": "customer_name", "operator": "contains", "value": "Ali"}],
        "column_order": ["customer_name", "follow_up_date"],
    }

    service.save(settings)

    assert json.loads(path.read_text(encoding="utf-8")) == settings
    assert service.load() == settings


def test_settings_service_saves_by_replacing_a_temp_file(tmp_path, monkeypatch):
    path = tmp_path / "daily_followup_report_settings.json"
    replacements = []
    original_replace = type(path).replace

    def recording_replace(self, target):
        replacements.append((self.name, type(path)(target).name))
        return original_replace(self, target)

    monkeypatch.setattr(type(path), "replace", recording_replace)

    DailyFollowupReportSettingsService(path).save({"selected_columns": ["notes"]})

    assert replacements == [
        ("daily_followup_report_settings.json.tmp", "daily_followup_report_settings.json")
    ]
    assert not path.with_suffix(".json.tmp").exists()


def test_settings_service_recovers_from_invalid_json(tmp_path):
    path = tmp_path / "daily_followup_report_settings.json"
    path.write_text("{not-json", encoding="utf-8")

    assert DailyFollowupReportSettingsService(path).load() == {}


def test_fetch_report_executes_parameterized_query_and_projects_selected_columns():
    review_service = _ReviewService(
        [{"daily_followup_id": 1, "customer_name": "Ali", "notes": "Call back"}]
    )
    service = DailyFollowupReportService(review_service)
    request = DailyFollowupReportRequest(
        selected_columns=["daily_followup_id", "customer_name"],
        filters=[SmartFilterCondition("notes", "contains", "Call")],
    )

    result = service.fetch_report(request)

    assert result.columns == service.validate_columns(["daily_followup_id", "customer_name"])
    assert result.rows == [{"daily_followup_id": 1, "customer_name": "Ali"}]
    assert "notes" not in result.rows[0]
    query, params = review_service.connection.calls[0]
    assert "notes ILIKE %s" in query
    assert params == ["%Call%"]


def test_load_filter_options_uses_existing_lookup_services():
    service = DailyFollowupReportService(_ReviewService())

    options = service.load_filter_options()

    assert options["statuses"] == [{"id": 1, "label": "Interested"}]
    assert options["employees"] == [{"id": 7, "label": "Ahmed"}]
    assert options["areas"] == [{"id": 3, "label": "Cairo"}]
    assert options["customers"] == [
        {"id": 1001, "label": "Ali - 1001", "name": "Ali", "code": 1001, "phone": "010"}
    ]


def test_customer_lookup_options_are_capped_and_labeled_for_dropdown():
    review_service = _ReviewService()
    service = DailyFollowupReportService(review_service)

    options = service.lookup_customers("  Ali  ", limit=500)

    assert review_service.customer_search_calls == [("Ali", 100)]
    assert options == [{"id": 1001, "label": "Ali - 1001", "name": "Ali", "code": 1001, "phone": "010"}]


def test_customer_name_filter_with_selected_customer_id_filters_by_id_not_display_text():
    service = DailyFollowupReportService()
    request = DailyFollowupReportRequest(
        selected_columns=["daily_followup_id"],
        filters=[SmartFilterCondition("customer_name", "equals", 500)],
    )

    query, params = service.build_query(request)

    where_sql = str(query).split(" WHERE ", 1)[1].split(" ORDER BY ", 1)[0]
    assert "d.customer_id = %s" in where_sql
    assert "c.customer_name" not in where_sql
    assert params == [500]


def test_excel_exporter_writes_only_selected_columns(tmp_path):
    service = DailyFollowupReportService()
    columns = service.validate_columns(["customer_name", "notes"])
    rows = [{"customer_name": "Ali", "notes": "Call back", "customer_phone": "010"}]
    path = tmp_path / "report.xlsx"

    ExcelExporter().export_table(path, "Daily Follow-up Smart Report", columns, rows)

    worksheet = load_workbook(path).active
    assert [worksheet.cell(1, 1).value, worksheet.cell(1, 2).value] == [
        columns[0].label,
        columns[1].label,
    ]
    assert [worksheet.cell(2, 1).value, worksheet.cell(2, 2).value] == ["Ali", "Call back"]
    assert worksheet.max_column == 2


def test_excel_exporter_sanitizes_invalid_sheet_title_characters(tmp_path):
    service = DailyFollowupReportService()
    columns = service.validate_columns(["customer_name"])
    path = tmp_path / "report.xlsx"

    ExcelExporter().export_table(path, "Daily/Follow-up:Smart*Report?", columns, [])

    assert load_workbook(path).active.title == "Daily Follow-up Smart Report"


def test_print_manager_html_uses_only_selected_columns():
    service = DailyFollowupReportService()
    columns = service.validate_columns(["customer_name", "notes"])
    rows = [{"customer_name": "Ali", "notes": "Call back", "customer_phone": "010"}]

    html = PrintManager().build_html("Daily Follow-up Smart Report", columns, rows)

    assert "Ali" in html
    assert "Call back" in html
    assert "010" not in html


def test_pdf_exporter_table_data_uses_only_selected_columns():
    service = DailyFollowupReportService()
    columns = service.validate_columns(["customer_name", "notes"])
    rows = [{"customer_name": "Ali", "notes": "Call back", "customer_phone": "010"}]

    table_data = PdfExporter().build_table_data(columns, rows)

    exporter = PdfExporter()
    assert table_data == [
        [exporter.format_text(columns[1].label), exporter.format_text(columns[0].label)],
        ["Call back", "Ali"],
    ]


def test_pdf_exporter_table_data_uses_rtl_column_order():
    service = DailyFollowupReportService()
    columns = service.validate_columns(["follow_up_date", "customer_name", "case_status_name"])
    rows = [
        {
            "follow_up_date": "2024-06-01",
            "customer_name": "Ali",
            "case_status_name": "Interested",
        }
    ]
    exporter = PdfExporter()

    table_data = exporter.build_table_data(columns, rows)

    assert table_data == [
        [
            exporter.format_text(columns[2].label),
            exporter.format_text(columns[1].label),
            exporter.format_text(columns[0].label),
        ],
        ["Interested", "Ali", "2024-06-01"],
    ]


def test_pdf_exporter_distributes_columns_across_available_page_width():
    service = DailyFollowupReportService()
    columns = service.validate_columns(["follow_up_date", "customer_name", "customer_phone", "case_status_name"])

    widths = PdfExporter().column_widths(columns, available_width=800)

    assert widths == [200, 200, 200, 200]
    assert sum(widths) == 800


def test_pdf_exporter_rejects_empty_selected_columns():
    with pytest.raises(ValueError, match="At least one PDF column"):
        PdfExporter().build_table_data([], [{"customer_name": "Ali"}])


def test_pdf_exporter_uses_arabic_capable_fonts_and_shapes_rtl_text(monkeypatch, tmp_path):
    exporter = PdfExporter()
    regular_font = tmp_path / "regular.ttf"
    bold_font = tmp_path / "bold.ttf"
    regular_font.touch()
    bold_font.touch()
    registered_fonts = []

    pdf_exporter_module._resolve_font_names.cache_clear()
    monkeypatch.setattr(pdf_exporter_module, "_font_candidates", lambda: [(regular_font, bold_font)])
    monkeypatch.setattr(
        pdf_exporter_module,
        "_register_font",
        lambda name, path: registered_fonts.append((name, path)),
    )

    try:
        font_names = exporter.resolve_font_names()
        arabic_label = "\u0627\u0633\u0645 \u0627\u0644\u0639\u0645\u064a\u0644"
        shaped_text = exporter.format_text(arabic_label)

        assert font_names.regular == "CRMArabic"
        assert font_names.bold == "CRMArabicBold"
        assert registered_fonts == [("CRMArabic", regular_font), ("CRMArabicBold", bold_font)]
        assert shaped_text != arabic_label
    finally:
        pdf_exporter_module._resolve_font_names.cache_clear()
