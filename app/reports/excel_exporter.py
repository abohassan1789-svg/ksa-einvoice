"""Excel export helpers."""

from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Write selected report columns to an XLSX workbook."""

    def export_table(
        self,
        path: str | Path,
        title: str,
        columns: list[Any],
        rows: list[dict[str, Any]],
    ) -> Path:
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self._safe_sheet_title(title)
        worksheet.append([column.label for column in columns])
        for row in rows:
            worksheet.append([self._cell_value(row.get(column.key)) for column in columns])
        header_fill = PatternFill("solid", fgColor="137A38")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for index, column in enumerate(columns, start=1):
            values = [str(column.label), *[str(row.get(column.key) or "") for row in rows]]
            width = min(max(len(value) for value in values) + 4, 45)
            worksheet.column_dimensions[get_column_letter(index)].width = width
        workbook.save(output_path)
        return output_path

    def export_with_summary(
        self,
        path: str | Path,
        title: str,
        subtitle_lines: list[str],
        columns: list[Any],
        rows: list[dict[str, Any]],
        summary_lines: list[str],
    ) -> Path:
        """Export a report sheet with a title, applied-filter lines, the data
        table and trailing summary totals. RTL sheet view for Arabic."""
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self._safe_sheet_title(title)
        worksheet.sheet_view.rightToLeft = True
        column_count = max(1, len(columns))

        def _append_merged(text: str, *, bold: bool, size: int, align: str, color: str | None = None) -> None:
            worksheet.append([text])
            row_index = worksheet.max_row
            worksheet.merge_cells(
                start_row=row_index, start_column=1, end_row=row_index, end_column=column_count
            )
            cell = worksheet.cell(row=row_index, column=1)
            cell.font = Font(bold=bold, size=size, color=color) if color else Font(bold=bold, size=size)
            cell.alignment = Alignment(horizontal=align, vertical="center")

        _append_merged(title, bold=True, size=15, align="center", color="137A38")
        for line in subtitle_lines:
            if line:
                _append_merged(line, bold=False, size=10, align="right")
        worksheet.append([])

        header_row_index = worksheet.max_row + 1
        worksheet.append([column.label for column in columns])
        header_fill = PatternFill("solid", fgColor="137A38")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[header_row_index]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            worksheet.append([self._cell_value(row.get(column.key)) for column in columns])

        worksheet.append([])
        for line in summary_lines:
            if line:
                _append_merged(line, bold=True, size=11, align="right")

        for index, column in enumerate(columns, start=1):
            values = [str(column.label), *[str(row.get(column.key) or "") for row in rows]]
            width = min(max(len(value) for value in values) + 4, 45)
            worksheet.column_dimensions[get_column_letter(index)].width = width
        workbook.save(output_path)
        return output_path

    @staticmethod
    def _cell_value(value: Any) -> Any:
        return "" if value is None else value

    @staticmethod
    def _safe_sheet_title(title: str) -> str:
        cleaned = re.sub(r"[\\*?:/\[\]]+", " ", str(title)).strip()
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned[:31] or "Report"
