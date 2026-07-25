"""Tests for the reusable Excel template import/export engine.

Covers the five requested behaviours: template column generation, missing-column
validation, invalid-row validation, duplicate prevention and saving valid rows
only. All tests use a fake data service, so no database is required.
"""

from __future__ import annotations

import datetime

import pytest
from openpyxl import load_workbook

from app.services.excel_io_service import (
    ExcelTemplateService,
    ExcelWorkbookError,
    io_spec_for,
)


class FakeDataService:
    """Minimal stand-in for ReviewDataService used by the Excel engine."""

    def __init__(self, existing=None, customers=None, statuses=None):
        # existing: {table_name: [row dicts]} for duplicate signatures.
        self.existing = existing or {}
        self.customers = customers or {}  # phone -> customer_id
        self.statuses = statuses or {}  # name -> case_status_id
        self.saved: list[tuple] = []  # (table_key, payload)

    def fetch_column_values(self, table_name, columns):
        return self.existing.get(table_name, [])

    def find_customer_by_phone(self, phone):
        cid = self.customers.get(str(phone).strip())
        return {"customer_id": cid} if cid is not None else None

    def find_case_status_by_name(self, name):
        sid = self.statuses.get(str(name).strip())
        return {"case_status_id": sid} if sid is not None else None

    def save_record(self, spec, payload, record_id):
        self.saved.append((spec.key, payload))
        return 1


# --- 1. Template column generation -----------------------------------------


def test_template_columns_match_customer_grid_exactly():
    io_spec = io_spec_for("customers")
    assert io_spec.headers == [
        "كود العميل",
        "اسم العميل",
        "الهاتف",
        "رقم المنطقة",
        "الوحدة",
    ]


def test_export_template_writes_headers_and_freezes_top_row(tmp_path):
    service = ExcelTemplateService(FakeDataService())
    io_spec = io_spec_for("employees")
    out = tmp_path / "employees_template.xlsx"

    service.export_template(io_spec, out)

    workbook = load_workbook(out)
    worksheet = workbook.active
    header = [cell.value for cell in worksheet[1]]
    assert header == io_spec.headers
    assert worksheet[1][0].font.bold is True
    assert worksheet.freeze_panes == "A2"
    # Template is empty apart from the header row.
    assert worksheet.max_row == 1


# --- 2. Missing / mismatched column validation -----------------------------


def test_check_headers_rejects_missing_column():
    service = ExcelTemplateService(FakeDataService())
    io_spec = io_spec_for("customers")
    header = ["كود العميل", "اسم العميل", "الهاتف", "رقم المنطقة"]  # missing "الوحدة"

    with pytest.raises(ExcelWorkbookError) as exc:
        service.check_headers(io_spec, header)
    assert "الوحدة" in str(exc.value)


def test_check_headers_accepts_exact_match_with_trailing_blanks():
    service = ExcelTemplateService(FakeDataService())
    io_spec = io_spec_for("case_statuses")
    # Trailing empty cells (added by some spreadsheet apps) are tolerated.
    service.check_headers(io_spec, [*io_spec.headers, "", None])


# --- 3. Invalid row validation ---------------------------------------------


def test_missing_required_field_marks_row_invalid():
    service = ExcelTemplateService(FakeDataService())
    io_spec = io_spec_for("customers")
    # Missing required customer name (col order: id, name, phone, area, unit).
    rows = service.build_import_rows(io_spec, [["", "", "0100000001", "5", "12"]])

    assert len(rows) == 1
    assert not rows[0].is_valid
    assert any("اسم العميل" in e for e in rows[0].errors)


def test_invalid_date_and_unknown_foreign_key_marked_invalid():
    data = FakeDataService(customers={"0100known": 1001}, statuses={"مهتم": 3})
    service = ExcelTemplateService(data)
    io_spec = io_spec_for("daily_followups")
    # cols: id, customer_name, customer_phone, follow_up_date, case_status_name, notes
    rows = service.build_import_rows(
        io_spec,
        [
            ["", "Ali", "0100known", "2026-13-40", "مهتم", "n"],  # bad date
            ["", "Sara", "0100unknown", "2026-07-06", "مهتم", "n"],  # bad phone
        ],
    )

    assert not rows[0].is_valid and any("تاريخ" in e for e in rows[0].errors)
    assert not rows[1].is_valid and any("رقم التليفون" in e or "غير معروفة" in e for e in rows[1].errors)


def test_valid_daily_followup_row_resolves_foreign_keys():
    data = FakeDataService(customers={"0100known": 1001}, statuses={"مهتم": 3})
    service = ExcelTemplateService(data)
    io_spec = io_spec_for("daily_followups")

    rows = service.build_import_rows(
        io_spec, [["", "Ali", "0100known", "2026-07-06", "مهتم", "note"]]
    )

    assert rows[0].is_valid
    assert rows[0].payload["customer_id"] == 1001
    assert rows[0].payload["case_status_id"] == 3
    assert rows[0].payload["follow_up_date"] == datetime.date(2026, 7, 6)
    assert rows[0].payload["notes"] == "note"


# --- 4. Duplicate prevention -----------------------------------------------


def test_duplicate_against_existing_database_row():
    data = FakeDataService(existing={"customers": [{"phone_number": "0100000001"}]})
    service = ExcelTemplateService(data)
    io_spec = io_spec_for("customers")

    rows = service.build_import_rows(io_spec, [["", "Ahmed", "0100000001", "5", "1"]])

    assert not rows[0].is_valid
    assert any("مكرر" in e for e in rows[0].errors)


def test_duplicate_within_the_same_file():
    service = ExcelTemplateService(FakeDataService())
    io_spec = io_spec_for("places")
    # places cols: place_id, place_number
    rows = service.build_import_rows(io_spec, [["", "A1"], ["", "A1"]])

    assert rows[0].is_valid
    assert not rows[1].is_valid
    assert any("مكرر" in e for e in rows[1].errors)


# --- 5. Save valid rows only ------------------------------------------------


def test_save_valid_rows_only_persists_valid_rows():
    data = FakeDataService()
    service = ExcelTemplateService(data)
    io_spec = io_spec_for("case_statuses")
    # case_statuses cols: case_status_id, case_status_name
    rows = service.build_import_rows(
        io_spec,
        [
            ["", "جديد"],  # valid
            ["", ""],  # invalid: missing required name
            ["", "مغلق"],  # valid
        ],
    )

    saved, failures = service.save_valid_rows(io_spec, rows)

    assert saved == 2
    assert failures == []
    saved_names = [payload.get("case_status_name") for _key, payload in data.saved]
    assert saved_names == ["جديد", "مغلق"]
